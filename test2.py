import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import re
import boto3
import json
from datetime import datetime
import os
import logging
from typing import Dict, List

logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s")
logger = logging.getLogger(__name__)

NOVA_PRO_MODEL_ID = "us.amazon.nova-pro-v1:0"
bedrock_client = boto3.client("bedrock-runtime", region_name="us-east-1")

def call_bedrock(prompt: str, max_tokens: int = 500) -> str:
    """Call AWS Bedrock Nova Pro"""
    try:
        request_body = {
            "schemaVersion": "messages-v1",
            "messages": [{"role": "user", "content": [{"text": prompt}]}],
            "inferenceConfig": {"maxTokens": max_tokens, "temperature": 0.5}
        }
        response = bedrock_client.invoke_model_with_response_stream(
            modelId=NOVA_PRO_MODEL_ID, body=json.dumps(request_body)
        )
        full_response = ""
        for event in response.get("body", []):
            chunk = event.get("chunk")
            if chunk:
                chunk_json = json.loads(chunk.get("bytes").decode())
                if content_block_delta := chunk_json.get("contentBlockDelta"):
                    if text := content_block_delta.get("delta", {}).get("text"):
                        full_response += text
        logger.info(f"Bedrock response length: {len(full_response)}")
        return full_response
    except Exception as e:
        logger.error(f"Bedrock API error: {e}")
        return ""

class CostReportAgent:
    def __init__(self, default_usd_to_inr: float, default_region: str = "US East (N. Virginia)"):
        self.usd_to_inr = default_usd_to_inr
        self.default_region = default_region
        logger.info("CostReportAgent initialized with USD to INR rate: %.2f", self.usd_to_inr)

    def extract_ec2_specs(self, instance_types: List[str]) -> Dict:
        if not instance_types:
            return {}

        prompt = f"""You are an AWS EC2 specifications expert. Provide EXACT official AWS specifications.

Instance types to lookup: {', '.join(instance_types)}

Return ONLY valid JSON (no markdown, no text):
{{
  "t3a.medium": {{"vCPUs": 2, "MemoryGiB": 4}},
  "m6a.large": {{"vCPUs": 2, "MemoryGiB": 8}},
  "c5.xlarge": {{"vCPUs": 4, "MemoryGiB": 8}}
}}

JSON:"""

        try:
            response = call_bedrock(prompt, max_tokens=600)
            logger.info(f"EC2 specs raw response: {response[:200]}")
            
            json_match = re.search(r'\{[^{}]*(?:\{[^{}]*\}[^{}]*)*\}', response, re.DOTALL)
            if json_match:
                json_str = json_match.group()
                ec2_specs = json.loads(json_str)
                logger.info(f"Successfully parsed EC2 specs: {ec2_specs}")
                
                for it in instance_types:
                    if it not in ec2_specs:
                        ec2_specs[it] = {"vCPUs": None, "MemoryGiB": None}
                return ec2_specs
            else:
                logger.warning("No JSON found in EC2 specs response")
        except json.JSONDecodeError as e:
            logger.error(f"JSON decode error for EC2 specs: {e}")
        except Exception as e:
            logger.error(f"Error extracting EC2 specs: {e}")
        
        return {it: {"vCPUs": None, "MemoryGiB": None} for it in instance_types}

    def extract_rds_specs(self, instance_types: List[str]) -> Dict:
        if not instance_types:
            return {}

        prompt = f"""You are an AWS RDS specifications expert. Provide EXACT official AWS RDS instance specifications.

RDS instance types to lookup: {', '.join(instance_types)}

Return ONLY valid JSON (no markdown, no text):
{{
  "db.t3.medium": {{"vCPUs": 2, "MemoryGiB": 4}},
  "db.m5.large": {{"vCPUs": 2, "MemoryGiB": 8}}
}}

JSON:"""

        try:
            response = call_bedrock(prompt, max_tokens=600)
            logger.info(f"RDS specs raw response: {response[:200]}")
            
            json_match = re.search(r'\{[^{}]*(?:\{[^{}]*\}[^{}]*)*\}', response, re.DOTALL)
            if json_match:
                json_str = json_match.group()
                rds_specs = json.loads(json_str)
                logger.info(f"Successfully parsed RDS specs: {rds_specs}")
                
                for it in instance_types:
                    if it not in rds_specs:
                        rds_specs[it] = {"vCPUs": None, "MemoryGiB": None}
                return rds_specs
            else:
                logger.warning("No JSON found in RDS specs response")
        except json.JSONDecodeError as e:
            logger.error(f"JSON decode error for RDS specs: {e}")
        except Exception as e:
            logger.error(f"Error extracting RDS specs: {e}")
        
        return {it: {"vCPUs": None, "MemoryGiB": None} for it in instance_types}

    def extract_service_config_summary(self, service_name: str, configuration_summary: str) -> str:
        """Extract ALL configuration details from the configuration summary"""
        try:
            config_summary = str(configuration_summary).strip()
            if not config_summary or config_summary == "":
                return f"{service_name}: Configuration details not available."
            
            # Extract all key-value pairs from the configuration
            extracted_info = []
            
            # Pattern to match: "Key (value)" or "Key: value (amount)"
            matches = re.findall(r"([A-Za-z][A-Za-z0-9\s:,.-]+?)\s*\(([^)]+)\)", config_summary)
            
            for key, value in matches:
                key = key.strip()
                value = value.strip()
                
                # Skip empty values, zeros, and "not selected" items
                if not value or value == "0" or "not selected" in value.lower():
                    continue
                
                # Clean up the key
                key_lower = key.lower()
                
                # Format the output based on the key type
                if "spice capacity" in key_lower:
                    extracted_info.append(f"{value} GB SPICE capacity")
                elif "number of authors" in key_lower:
                    extracted_info.append(f"{value} Authors")
                elif "number of readers" in key_lower and "pro" not in key_lower:
                    extracted_info.append(f"{value} Readers")
                elif "number of reader pros" in key_lower:
                    extracted_info.append(f"{value} Reader Pros")
                elif "requests per minute" in key_lower:
                    extracted_info.append(f"{value} requests per minute")
                elif "hours per day" in key_lower:
                    extracted_info.append(f"{value} hours per day")
                elif "input tokens per request" in key_lower:
                    extracted_info.append(f"{value} input tokens per request")
                elif "output tokens per request" in key_lower:
                    extracted_info.append(f"{value} output tokens per request")
                elif "input images per request" in key_lower:
                    extracted_info.append(f"{value} input images per request")
                elif "input image length" in key_lower or "input image width" in key_lower:
                    extracted_info.append(f"{value} pixels ({key.split('(')[-1].strip()})")
                elif "requests per batch" in key_lower:
                    extracted_info.append(f"{value} requests per batch")
                elif "number of documents" in key_lower:
                    extracted_info.append(f"{value} documents per month")
                elif "tokens per document" in key_lower:
                    extracted_info.append(f"{value} tokens per document")
                elif "number of records" in key_lower:
                    extracted_info.append(f"{value} records")
                elif "tokens per record" in key_lower:
                    if "input" in key_lower:
                        extracted_info.append(f"{value} input tokens per record")
                    elif "output" in key_lower:
                        extracted_info.append(f"{value} output tokens per record")
                    else:
                        extracted_info.append(f"{value} tokens per record")
                elif "storage" in key_lower:
                    extracted_info.append(f"{value} storage")
                elif "number of requests" in key_lower:
                    extracted_info.append(f"{value} requests")
                elif "concurrency" in key_lower:
                    extracted_info.append(f"{value} concurrency")
                elif "ephemeral storage" in key_lower:
                    extracted_info.append(f"{value} ephemeral storage")
                elif "number of pages" in key_lower:
                    extracted_info.append(f"{value} pages")
                elif "standard queue" in key_lower:
                    extracted_info.append(f"{value} standard queue requests")
                elif "fifo queue" in key_lower:
                    extracted_info.append(f"{value} FIFO queue requests")
                elif "fair queue" in key_lower:
                    extracted_info.append(f"{value} fair queue requests")
                elif "dt inbound" in key_lower or "inbound" in key_lower:
                    if value != "0 TB per month" and value != "0 GB per month":
                        extracted_info.append(f"{value} inbound data transfer")
                elif "dt outbound" in key_lower or "outbound" in key_lower:
                    if value != "0 TB per month" and value != "0 GB per month":
                        extracted_info.append(f"{value} outbound data transfer")
                elif "dt intra-region" in key_lower:
                    if value != "0 TB per month" and value != "0 GB per month":
                        extracted_info.append(f"{value} intra-region data transfer")
                elif "s3 standard" in key_lower:
                    extracted_info.append(f"{value} S3 standard storage")
                elif "architecture" in key_lower:
                    extracted_info.append(f"{value} architecture")
                elif "invoke mode" in key_lower:
                    extracted_info.append(f"{value} invoke mode")
                elif "inference route" in key_lower:
                    extracted_info.append(f"{value} inference route")
                elif "inference type" in key_lower:
                    extracted_info.append(f"{value} inference type")
                else:
                    # Generic format for any other configuration
                    extracted_info.append(f"{value} {key.lower()}")
            
            if extracted_info:
                return f"{service_name}: {', '.join(extracted_info)} considered"
            else:
                return f"{service_name}: Configuration as per requirements"
                
        except Exception as e:
            logger.error(f"Error extracting config summary for {service_name}: {e}")
            return f"{service_name}: Configuration as per requirements"

    def generate_best_practices(self, services: List[str]) -> List[str]:
        if not services:
            return ["No specific services detected. General AWS best practices apply."]

        services_str = ", ".join(services[:10])
        prompt = f"""You are an AWS Solutions Architect. Based on these AWS services: {services_str}

Provide 5 specific, actionable best practice recommendations focusing on:
- Cost optimization
- Security
- Performance
- Operational excellence

Format: Number each recommendation 1-5. Keep each to 1-2 sentences. Be specific to the services mentioned.

Start with "1." immediately:"""

        try:
            response = call_bedrock(prompt, max_tokens=500)
            if response:
                logger.info(f"Best practices response: {response[:100]}")
                lines = []
                for line in response.split('\n'):
                    line = line.strip()
                    if line and re.match(r'^\d+\.', line):
                        lines.append(line)
                
                if lines:
                    logger.info(f"Found {len(lines)} best practice lines")
                    return lines[:5]
                else:
                    logger.warning("No numbered lines found in response")
            else:
                logger.warning("Empty response for best practices")
        except Exception as e:
            logger.error(f"Error generating best practices: {e}")
        
        return [
            "1. Implement IAM roles with least privilege principles for enhanced security.",
            "2. Enable CloudTrail and CloudWatch for comprehensive auditing and monitoring.",
            "3. Use AWS Cost Explorer and set up billing alerts for cost optimization.",
            "4. Implement auto-scaling and right-sizing for compute resources.",
            "5. Enable encryption at rest and in transit for all sensitive data."
        ]

    def extract_ec2_values(self, configuration_summary: str) -> tuple:
        try:
            os_match = re.search(r"operating\s*system\s*\((.*?)\)", configuration_summary, re.I)
            type_match = re.search(r"ec2\s*instance\s*\((.*?)\)", configuration_summary, re.I)
            price_match = re.search(r"pricing\s*strategy\s*\((.*?)\)", configuration_summary, re.I)
            return (
                os_match.group(1).strip() if os_match else None,
                type_match.group(1).strip() if type_match else None,
                price_match.group(1).strip() if price_match else None,
            )
        except:
            return None, None, None

    def extract_rds_values(self, configuration_summary: str, service_name: str) -> tuple:
        try:
            db_type = None
            db_engines = ["MySQL", "PostgreSQL", "MariaDB", "Oracle", "SQL Server", "Aurora"]
            for engine in db_engines:
                if engine.upper() in service_name.upper():
                    db_type = engine
                    break
            
            instance_type = None
            type_match = re.search(r"(?:rds\s*instance|instance\s*type|instance)\s*\((.*?)\)", configuration_summary, re.I)
            if type_match:
                instance_type = type_match.group(1).strip()
            else:
                type_match = re.search(r"(db\.[a-z0-9]+\.[a-z0-9]+)", configuration_summary, re.I)
                if type_match:
                    instance_type = type_match.group(1).strip()
            
            pricing_model = None
            price_match = re.search(r"(?:pricing\s*strategy|reserved|upfront)\s*\((.*?)\)", configuration_summary, re.I)
            if price_match:
                pricing_model = price_match.group(1).strip()
            else:
                if "reserved" in configuration_summary.lower():
                    if "no upfront" in configuration_summary.lower():
                        pricing_model = "Reserved No Upfront"
                    else:
                        pricing_model = "Reserved"
            
            return (db_type, instance_type, pricing_model)
        except Exception as e:
            logger.error(f"Error extracting RDS values: {e}")
            return None, None, None

    def generate_cost_report(self, input_file: str, output_file: str, customer_name: str,
                            usd_to_inr: float, region: str, pricing_link: str = ""):
        os.makedirs(os.path.dirname(output_file) or '.', exist_ok=True)

        header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        total_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        pricing_link_fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
        best_practice_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
        note_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        thin = Side(style='thin')
        full_border = Border(left=thin, right=thin, top=thin, bottom=thin)

        try:
            if not os.path.exists(input_file):
                raise FileNotFoundError(f"Input CSV not found: {input_file}")

            df = pd.read_csv(input_file, skiprows=7)
            df.columns = df.columns.str.lower().str.strip()

            service_col = next((c for c in df.columns if "service" in c), None)
            monthly_col = next((c for c in df.columns if "monthly" in c or "cost" in c), None)
            config_col = next((c for c in df.columns if "configuration" in c or "summary" in c or "config" in c), None)

            if not all([service_col, monthly_col, config_col]):
                raise ValueError("Required columns not found")

            data = df[[service_col, monthly_col, config_col]].fillna("")
            
            # Get column positions for safe indexing
            service_idx = 0  # First column in data
            monthly_idx = 1  # Second column in data  
            config_idx = 2   # Third column in data

            has_ec2 = any("EC2" in str(data.iloc[i, service_idx]).upper() for i in range(len(data)))
            has_rds = any("RDS" in str(data.iloc[i, service_idx]).upper() for i in range(len(data)))

            instance_types = set()
            ec2_specs = {}
            if has_ec2:
                for val in data[config_col]:
                    if "EC2" in str(val).upper():
                        m = re.search(r"ec2\s*instance\s*\((.*?)\)", str(val), re.I)
                        if m:
                            instance_types.add(m.group(1).strip())
                ec2_specs = self.extract_ec2_specs(list(instance_types))

            rds_instance_types = set()
            rds_specs = {}
            if has_rds:
                for i in range(len(data)):
                    svc_name = str(data.iloc[i, service_idx]).upper()
                    if "RDS" in svc_name or "MYSQL" in svc_name or "POSTGRESQL" in svc_name or "MARIADB" in svc_name or "ORACLE" in svc_name or "SQL SERVER" in svc_name:
                        config_val = str(data.iloc[i, config_idx])
                        # Look for instance type patterns
                        m = re.search(r"(db\.[a-z0-9]+\.[a-z0-9]+)", config_val, re.I)
                        if m:
                            rds_instance_types.add(m.group(1).strip())
                            logger.info(f"Found RDS instance type: {m.group(1).strip()}")
                if rds_instance_types:
                    logger.info(f"Extracting specs for RDS instances: {rds_instance_types}")
                    rds_specs = self.extract_rds_specs(list(rds_instance_types))

            specs_failed = False
            if ec2_specs:
                specs_failed = any(v.get("vCPUs") is None and v.get("MemoryGiB") is None for v in ec2_specs.values() if isinstance(v, dict))
            
            rds_specs_failed = False
            if rds_specs:
                rds_specs_failed = any(v.get("vCPUs") is None and v.get("MemoryGiB") is None for v in rds_specs.values() if isinstance(v, dict))

            seen_services = set()
            cleaned_services = []
            for i in range(len(data)):
                svc = str(data.iloc[i, service_idx]).strip()
                if svc and svc not in seen_services:
                    seen_services.add(svc)
                    clean_name = svc.split('(')[0].strip()
                    cleaned_services.append(clean_name)

            wb = openpyxl.Workbook()
            summary = wb.active
            summary.title = "Summary"

            summary.merge_cells("A1:B1")
            summary["A1"] = "Cost Estimation Summary"
            summary["A1"].font = Font(bold=True, color="000000")
            summary["A1"].alignment = Alignment(horizontal="center")
            summary["A1"].fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

            summary["A2"] = "Description"
            summary["B2"] = "Monthly Cost"
            for cell in [summary["A2"], summary["B2"]]:
                cell.fill = PatternFill(start_color="A7C5EB", end_color="A7C5EB", fill_type="solid")
                cell.border = full_border
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")

            total_usd = data[monthly_col].apply(pd.to_numeric, errors='coerce').sum()
            total_inr = total_usd * usd_to_inr

            summary["A3"] = "AWS Resource Cost (without TAX)"
            # B3 will be set after total_r is known (cross-sheet reference)
            summary["B3"].number_format = "₹#,##0.00"
            summary["B3"].alignment = Alignment(horizontal="center")

            for r in [3]:
                for c in [1, 2]:
                    summary.cell(r, c).border = full_border

            summary.column_dimensions["A"].width = 30
            summary.column_dimensions["B"].width = 15

            sheet = wb.create_sheet("AWS Services")
            sheet.merge_cells("A1:K1" if (has_ec2 or has_rds) else "A1:E1")
            sheet["A1"] = f"Cost Estimation Report For {customer_name}"
            sheet["A1"].font = Font(bold=True, color="000000")
            sheet["A1"].alignment = Alignment(horizontal="center")
            sheet["A1"].fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

            if has_ec2 or has_rds:
                headers = [
                    "S.NO", "Instance Type", "vCPU", "RAM", "Operating System/Database",
                    "Running Hours", "Pricing Model", "Services", "Per Month USD", "Per Month INR", "Per Year INR"
                ]
                usd_col = 9
                inr_col = 10
                yearly_col = 11
            else:
                headers = ["S.NO", "Services", "Per Month USD", "Per Month INR", "Per Year INR"]
                usd_col = 3
                inr_col = 4
                yearly_col = 5

            for col, hdr in enumerate(headers, 1):
                cell = sheet.cell(2, col, hdr)
                cell.fill = header_fill
                cell.border = full_border
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            row = 3
            counter = 1
            rate_str = f"{usd_to_inr:.4f}"
            for i in range(len(data)):
                full_service = str(data.iloc[i, service_idx]).strip()
                if not full_service:
                    continue
                try:
                    usd = float(data.iloc[i, monthly_idx])
                except:
                    continue

                is_ec2 = "EC2" in full_service.upper()
                is_rds = "RDS" in full_service.upper() or "MYSQL" in full_service.upper() or "POSTGRESQL" in full_service.upper() or "MARIADB" in full_service.upper() or "ORACLE" in full_service.upper() or "SQL SERVER" in full_service.upper()

                if (has_ec2 or has_rds) and (is_ec2 or is_rds):
                    config_val = str(data.iloc[i, config_idx])
                    if is_ec2:
                        os_val, instance_type, price_model = self.extract_ec2_values(config_val)
                        spec = ec2_specs.get(instance_type or "", {"vCPUs": None, "MemoryGiB": None})
                    else:  # is_rds
                        os_val, instance_type, price_model = self.extract_rds_values(config_val, full_service)
                        spec = rds_specs.get(instance_type or "", {"vCPUs": None, "MemoryGiB": None})
                    
                    # Ensure spec is a dict
                    if not isinstance(spec, dict):
                        spec = {"vCPUs": None, "MemoryGiB": None}

                    values = [
                        (1, counter), (2, instance_type or ""), (3, spec.get('vCPUs')),
                        (4, spec.get('MemoryGiB')), (5, os_val or ""), (6, "730 hours"),
                        (7, price_model or ""), (8, full_service), (9, usd)
                    ]

                    for col, val in values:
                        cell = sheet.cell(row, col, val)
                        cell.border = full_border
                        if col == 9:
                            cell.number_format = '$#,##0.00'
                        align_h = "right" if col in [2, 8] else ("left" if col < 9 else "right")
                        cell.alignment = Alignment(horizontal=align_h, vertical="center")

                    for col in [10, 11]:
                        formula = f"=I{row}*{rate_str}" if col == 10 else f"=J{row}*12"
                        cell = sheet.cell(row, col, value=formula)
                        cell.number_format = "₹#,##0.00"
                        cell.border = full_border
                        cell.alignment = Alignment(horizontal="right", vertical="center")
                else:
                    sheet.cell(row, 1, counter).border = full_border
                    sheet.cell(row, 1).alignment = Alignment(horizontal="right", vertical="center")

                    merge_end = usd_col - 1
                    sheet.merge_cells(f"B{row}:{openpyxl.utils.get_column_letter(merge_end)}{row}")
                    cell = sheet.cell(row, 2, full_service)
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                    cell.border = full_border

                    for c in range(2, usd_col):
                        sheet.cell(row, c).border = full_border

                    sheet.cell(row, usd_col, usd).number_format = '$#,##0.00'
                    sheet.cell(row, usd_col).border = full_border
                    sheet.cell(row, usd_col).alignment = Alignment(horizontal="right", vertical="center")

                    sheet.cell(row, inr_col, value=f"={openpyxl.utils.get_column_letter(usd_col)}{row}*{rate_str}").number_format = "₹#,##0.00"
                    sheet.cell(row, inr_col).border = full_border
                    sheet.cell(row, inr_col).alignment = Alignment(horizontal="right", vertical="center")
                    sheet.cell(row, yearly_col, value=f"={openpyxl.utils.get_column_letter(inr_col)}{row}*12").number_format = "₹#,##0.00"
                    sheet.cell(row, yearly_col).border = full_border
                    sheet.cell(row, yearly_col).alignment = Alignment(horizontal="right", vertical="center")

                counter += 1
                row += 1

            last_data_row = row - 1

            total_r = row
            merge_end_total = "H" if (has_ec2 or has_rds) else openpyxl.utils.get_column_letter(usd_col - 1)
            sheet.merge_cells(f"A{total_r}:{merge_end_total}{total_r}")
            sheet.cell(total_r, 1, "Total Cost").alignment = Alignment(horizontal="right")
            sheet.cell(total_r, 1).border = full_border

            for c in range(1, (12 if (has_ec2 or has_rds) else usd_col + 3)):
                sheet.cell(total_r, c).border = full_border
                sheet.cell(total_r, c).fill = total_fill

            if last_data_row >= 3:
                sheet.cell(total_r, usd_col).value = f"=SUM({openpyxl.utils.get_column_letter(usd_col)}3:{openpyxl.utils.get_column_letter(usd_col)}{last_data_row})"
                sheet.cell(total_r, inr_col).value = f"=SUM({openpyxl.utils.get_column_letter(inr_col)}3:{openpyxl.utils.get_column_letter(inr_col)}{last_data_row})"
                sheet.cell(total_r, yearly_col).value = f"=SUM({openpyxl.utils.get_column_letter(yearly_col)}3:{openpyxl.utils.get_column_letter(yearly_col)}{last_data_row})"
            else:
                sheet.cell(total_r, usd_col).value = 0
                sheet.cell(total_r, inr_col).value = 0
                sheet.cell(total_r, yearly_col).value = 0

            sheet.cell(total_r, usd_col).number_format = '$#,##0.00'
            sheet.cell(total_r, inr_col).number_format = '₹#,##0.00'
            sheet.cell(total_r, yearly_col).number_format = '₹#,##0.00'

            # Now set Summary B3 as cross-sheet formula referencing total INR cell in AWS Services
            inr_col_letter = openpyxl.utils.get_column_letter(inr_col)
            summary["B3"] = f"='AWS Services'!{inr_col_letter}{total_r}"

            pl_row = total_r + 1
            merge_end_pl = "H" if (has_ec2 or has_rds) else openpyxl.utils.get_column_letter(usd_col - 1)
            sheet.merge_cells(f"A{pl_row}:{merge_end_pl}{pl_row}")
            sheet.cell(pl_row, 1, "Pricing Link").alignment = Alignment(horizontal="right")

            for c in range(1, (12 if (has_ec2 or has_rds) else usd_col + 3)):
                sheet.cell(pl_row, c).border = full_border
                sheet.cell(pl_row, c).fill = pricing_link_fill

            sheet.cell(pl_row, usd_col if not (has_ec2 or has_rds) else 9, pricing_link or "Not provided")

            note_row = pl_row + 1
            note_title = sheet.cell(note_row, 1, "Note:")
            note_title.fill = note_fill

            notes = [
                "1. The given costs are considered as estimation based on the Monthly usage actual amount will get vary.",
                f"2. {region} region is considered for this workload.",
                f"3. The exchange rate is considered as ₹{usd_to_inr:.2f} as per {datetime.now().strftime('%d/%m/%y')} date."
            ]

            for i, note_text in enumerate(notes, 1):
                sheet.cell(note_row + i, 1, note_text)

            cur_note_row = note_row + len(notes) + 1
            note_sno = 4

            if specs_failed and has_ec2:
                sheet.cell(cur_note_row, 1, f"{note_sno}. Failed to extract EC2 specs (vCPU, RAM) from model.")
                cur_note_row += 1
                note_sno += 1

            if rds_specs_failed and has_rds:
                sheet.cell(cur_note_row, 1, f"{note_sno}. Failed to extract RDS specs (vCPU, RAM) from model.")
                cur_note_row += 1
                note_sno += 1

            seen = set()
            for i in range(len(data)):
                svc = str(data.iloc[i, service_idx]).strip()
                if svc and svc not in seen:
                    seen.add(svc)
                    clean = svc.split('(')[0].strip()
                    config_val = str(data.iloc[i, config_idx])
                    desc = self.extract_service_config_summary(clean, config_val)
                    sheet.cell(cur_note_row, 1, f"{note_sno}. {desc}")
                    cur_note_row += 1
                    note_sno += 1

            bp_row = cur_note_row
            bp_title = sheet.cell(bp_row, 1, "Best Practices:")
            bp_title.fill = best_practice_fill

            best_practices = self.generate_best_practices(cleaned_services)

            for i, line in enumerate(best_practices, 1):
                sheet.cell(bp_row + i, 1, line)

            if has_ec2 or has_rds:
                widths = [6, 28, 10, 10, 18, 14, 16, 40, 14, 14, 14]
            else:
                widths = [6, 50, 16, 16, 16]

            for col, width in enumerate(widths, 1):
                sheet.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

            wb.save(output_file)
            logger.info(f"Report saved: {output_file}")
            return {"status": "success", "file": output_file}

        except Exception as e:
            logger.exception("Report generation failed")
            return {"status": "error", "message": str(e)}